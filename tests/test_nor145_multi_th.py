"""Tests for Nor145MultipleTHParser."""

import datetime as dt

import pandas as pd
import pytest

from pycoustic import Log
from pycoustic.parsers.nor145_multi_th import (
    DIRECT_BROADBAND_MAP,
    SPECTRAL_FAMILY_MAP,
    Nor145MultipleTHParser,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def nor145_path() -> str:
    import os
    path = os.path.join(
        os.path.dirname(__file__),
        "nor145-multiple-th.xlsx",
    )
    if not os.path.exists(path):
        pytest.skip("nor145-multiple-th.xlsx not found")
    return path


@pytest.fixture(scope="module")
def parsed_results(nor145_path: str) -> list[tuple[str, pd.DataFrame]]:
    parser = Nor145MultipleTHParser()
    return parser.parse_all(nor145_path)


# ---------------------------------------------------------------------------
# can_parse
# ---------------------------------------------------------------------------

def test_can_parse_nor145(nor145_path: str) -> None:
    assert Nor145MultipleTHParser.can_parse(nor145_path)


def test_cannot_parse_csv(tmp_path) -> None:
    csv = tmp_path / "test.csv"
    csv.write_text("Time,Leq A\n2024/01/01 12:00,50\n")
    assert not Nor145MultipleTHParser.can_parse(str(csv))


# ---------------------------------------------------------------------------
# parse_all
# ---------------------------------------------------------------------------

def test_parse_all_yields_21_histories(parsed_results: list) -> None:
    assert len(parsed_results) == 21


def test_all_results_have_datetime_index(parsed_results: list) -> None:
    for name, df in parsed_results:
        assert isinstance(df.index, pd.DatetimeIndex), f"{name}: index is {type(df.index)}"


def test_all_results_have_broadband_columns(parsed_results: list) -> None:
    for name, df in parsed_results:
        assert "Leq A" in df.columns, f"{name}: missing Leq A"
        assert "Lmax A" in df.columns, f"{name}: missing Lmax A"


def test_all_results_have_spectral_columns(parsed_results: list) -> None:
    for name, df in parsed_results:
        assert "Leq 63" in df.columns, f"{name}: missing Leq 63"
        assert "Lmax 125" in df.columns, f"{name}: missing Lmax 125"
        assert "Leq 1000" in df.columns, f"{name}: missing Leq 1000"


def test_first_result_column_count(parsed_results: list) -> None:
    _, df = parsed_results[0]
    # 12 broadband + 31 spectral Leq + 31 spectral Lmax + 31 spectral Lmin = 105
    assert df.shape[1] == 105


def test_measurement_names_match_overview(parsed_results: list) -> None:
    expected = {
        "grantham L1 1",
        "grantham l2 1",
        "grantham l1 2",
        "grantham l2 2",
        "grantham l1 3",
        "grantham l2 3",
        "A to R1",
        "r1 bg",
        "R2 bg office",
        "B R2",
        "B R2 ft2",
        "B R2 FT3",
        "C R2 ft1",
        "C R3 FT1",
        "C R2 Ft2 mostly airborne",
        "C R3 FT2",
        "D R3 FT1",
        "D R3 FT 2",
        "B R2 FT1",
        "B R2 FT 2",
        "B R2 FT 4",
    }
    actual = {name for name, _ in parsed_results}
    assert actual == expected


def test_all_results_non_empty(parsed_results: list) -> None:
    for name, df in parsed_results:
        assert not df.empty, f"{name}: empty DataFrame"
        assert df.shape[0] > 0, f"{name}: {df.shape[0]} rows"


def test_parse_raises_not_implemented_error(nor145_path: str) -> None:
    parser = Nor145MultipleTHParser()
    with pytest.raises(NotImplementedError):
        parser.parse(nor145_path)


# ---------------------------------------------------------------------------
# Log.from_dataframe integration
# ---------------------------------------------------------------------------

def test_from_dataframe_creates_valid_log(parsed_results: list) -> None:
    name, df = parsed_results[0]
    log = Log.from_dataframe(df, filepath="test.xlsx", name=name)
    assert log.get_start() is not None
    assert log.get_end() is not None
    assert log.get_data() is not None
    assert not log.get_data().empty
    # Verify multi-index columns are set up
    assert isinstance(log.get_data().columns, pd.MultiIndex)


def test_from_dataframe_preserves_columns(parsed_results: list) -> None:
    name, df = parsed_results[0]
    log = Log.from_dataframe(df, filepath="test.xlsx", name=name)
    data = log.get_data()
    # Check A-weighted columns exist as tuples
    assert ("Leq", "A") in data.columns
    assert ("Lmax", "A") in data.columns
    # Check spectral column
    assert ("Leq", 31.5) in data.columns or ("Leq", 31.5) in data.columns


# ---------------------------------------------------------------------------
# Column mapping helpers
# ---------------------------------------------------------------------------

def test_broadband_map_covers_all_expected() -> None:
    expected_entries = 12  # A/C/Z × 4 metrics (Leq, Lmax, Lmin, Lpeak)
    assert len(DIRECT_BROADBAND_MAP) == expected_entries


def test_spectral_family_map_covers_lef_lffmax_lffmin() -> None:
    assert SPECTRAL_FAMILY_MAP["Lfeq"] == "Leq"
    assert SPECTRAL_FAMILY_MAP["LfFmax"] == "Lmax"
    assert SPECTRAL_FAMILY_MAP["LfFmin"] == "Lmin"


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

def test_non_existent_file() -> None:
    parser = Nor145MultipleTHParser()
    with pytest.raises(FileNotFoundError):
        parser.parse_all("nonexistent.xlsx")


def test_cannot_parse_empty_xlsx(tmp_path) -> None:
    import openpyxl
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(path))
    assert not Nor145MultipleTHParser.can_parse(str(path))


# ---------------------------------------------------------------------------
# File without overview sheets (only profile / time-history sheets)
# ---------------------------------------------------------------------------

def _write_minimal_profile_xlsx(path: str, measurements: list[tuple[str, int]]) -> None:
    """Write a minimal Nor145-style XLSX with only profile sheets.

    Each *measurement* is ``(name, num_rows)``.  The resulting workbook has
    **no** overview sheet — just the profile sheets a Nor145 export would
    produce for each measurement.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    profile_headers = [
        None,                          # col 0 — measurement-name placeholder
        None,                          # col 1 — timestamp
        "Markers",
        "LAeq [dB]", "LAFmax [dB]", "LAFmin [dB]", "LApeak [dB]",
        "LCeq [dB]", "LCFmax [dB]", "LCFmin [dB]", "LCpeak [dB]",
        "LZeq [dB]", "LZFmax [dB]", "LZFmin [dB]", "LZpeak [dB]",
        "Lfeq 63 Hz  (1/3) [dB]",
        "LfFmax 63 Hz  (1/3) [dB]",
        "LfFmin 63 Hz  (1/3) [dB]",
    ]

    for name, num_rows in measurements:
        ws = wb.create_sheet(title=name[:31])
        # Row 0: measurement name
        ws.cell(row=1, column=1, value=name)
        # Row 1: column headers
        for col_idx, header in enumerate(profile_headers, start=1):
            if header is not None:
                ws.cell(row=2, column=col_idx, value=header)
        # Rows 2+: data
        base = pd.Timestamp("2026-06-08 10:00:00")
        for r in range(num_rows):
            ts = base + pd.Timedelta(milliseconds=250 * r)
            ws.cell(row=3 + r, column=1, value=r)          # counter
            ws.cell(row=3 + r, column=2, value=ts)          # timestamp
            ws.cell(row=3 + r, column=3, value="Battery")   # marker
            ws.cell(row=3 + r, column=4, value=50.0)        # LAeq
            ws.cell(row=3 + r, column=5, value=55.0)        # LAFmax
            ws.cell(row=3 + r, column=6, value=45.0)        # LAFmin
            ws.cell(row=3 + r, column=7, value=80.0)        # LApeak
            ws.cell(row=3 + r, column=8, value=53.0)        # LCeq
            ws.cell(row=3 + r, column=9, value=58.0)        # LCFmax
            ws.cell(row=3 + r, column=10, value=48.0)       # LCFmin
            ws.cell(row=3 + r, column=11, value=83.0)       # LCpeak
            ws.cell(row=3 + r, column=12, value=54.0)       # LZeq
            ws.cell(row=3 + r, column=13, value=59.0)       # LZFmax
            ws.cell(row=3 + r, column=14, value=49.0)       # LZFmin
            ws.cell(row=3 + r, column=15, value=85.0)       # LZpeak
            ws.cell(row=3 + r, column=16, value=30.0)       # Leq 63
            ws.cell(row=3 + r, column=17, value=35.0)       # Lmax 63
            ws.cell(row=3 + r, column=18, value=25.0)       # Lmin 63

    wb.save(path)


def test_can_parse_without_overview_sheets(tmp_path) -> None:
    """can_parse should accept a file that contains only profile sheets."""
    path = tmp_path / "no_overview.xlsx"
    _write_minimal_profile_xlsx(str(path), [("Pos A", 5), ("Pos B", 3)])
    assert Nor145MultipleTHParser.can_parse(str(path))


def test_can_parse_single_profile_no_overview(tmp_path) -> None:
    """A single profile sheet alone (no overview) should still be recognised."""
    path = tmp_path / "single_profile.xlsx"
    _write_minimal_profile_xlsx(str(path), [("Pos A", 4)])
    assert Nor145MultipleTHParser.can_parse(str(path))


def test_parse_all_without_overview_sheets(tmp_path) -> None:
    """parse_all should work when no overview sheets are present."""
    path = tmp_path / "no_overview.xlsx"
    _write_minimal_profile_xlsx(str(path), [("Pos A", 5), ("Pos B", 3)])

    parser = Nor145MultipleTHParser()
    results = parser.parse_all(str(path))
    assert len(results) == 2

    names = {name for name, _ in results}
    assert names == {"Pos A", "Pos B"}

    for name, df in results:
        assert isinstance(df.index, pd.DatetimeIndex)
        assert "Leq A" in df.columns
        assert "Lmax A" in df.columns
        assert "Leq 63" in df.columns
        assert not df.empty


def test_parse_all_single_profile_no_overview(tmp_path) -> None:
    """A single profile sheet alone should parse successfully."""
    path = tmp_path / "single_profile.xlsx"
    _write_minimal_profile_xlsx(str(path), [("Pos A", 4)])

    parser = Nor145MultipleTHParser()
    results = parser.parse_all(str(path))
    assert len(results) == 1
    assert results[0][0] == "Pos A"
