"""Oracle test for Log.get_nth_high_low using pre-computed nth-highest/lowest values."""

import os
import datetime as dt

import openpyxl
import pandas as pd
import pytest

from pycoustic import Log
from pycoustic.parsers.nor145_multi_th import Nor145MultipleTHParser


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _oracle_col_to_pycoustic(col_name: str) -> tuple:
    """Convert an oracle column header like 'Lmax 20' or 'Leq A' to a
    pycoustic MultiIndex tuple like ('Lmax', 20.0) or ('Leq', 'A')."""
    parts = col_name.split(" ", 1)
    metric = parts[0]
    label = parts[1]
    # Try numeric conversion — spectral bands are floats in pycoustic
    try:
        label = float(label)
    except ValueError:
        pass  # keep as string (e.g. "A", "C")
    return (metric, label)


def _build_sheet_name_to_meas_name(nor145_path: str) -> dict[str, str]:
    """Map XLSX sheet names to the measurement names the parser extracts."""
    import openpyxl as _xl
    wb = _xl.load_workbook(nor145_path, read_only=True)
    mapping = {}
    for sheet_name in wb.sheetnames:
        raw = pd.read_excel(nor145_path, sheet_name=sheet_name, header=None)
        meas_name = str(raw.iloc[0, 0]).strip()
        mapping[sheet_name] = meas_name
    wb.close()
    return mapping


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def nor145_path() -> str:
    path = os.path.join(os.path.dirname(__file__), "nor145-multiple-th.xlsx")
    if not os.path.exists(path):
        pytest.skip("nor145-multiple-th.xlsx not found")
    return path


@pytest.fixture(scope="module")
def oracle_path() -> str:
    path = os.path.join(os.path.dirname(__file__), "nth-highest-lowest-oracle.xlsx")
    if not os.path.exists(path):
        pytest.skip("nth-highest-lowest-oracle.xlsx not found")
    return path


@pytest.fixture(scope="module")
def name_to_log(nor145_path: str) -> dict[str, Log]:
    """Parse all Nor145 profiles and return {measurement_name: Log}."""
    parser = Nor145MultipleTHParser()
    results = parser.parse_all(nor145_path)
    return {name: Log.from_dataframe(df, name=name) for name, df in results}


@pytest.fixture(scope="module")
def sheet_to_name(nor145_path: str) -> dict[str, str]:
    """Return {XLSX_sheet_name: measurement_name} mapping."""
    return _build_sheet_name_to_meas_name(nor145_path)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_nth_high_low_oracle_high(oracle_path, name_to_log, sheet_to_name):
    """Compare get_nth_high_low output against pre-computed nth-highest Lmax values."""
    wb = openpyxl.load_workbook(oracle_path)
    ws = wb["high"]

    # Read header row to get spectral column names
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    spectral_col_names = header[3:]  # skip "tab name", "timestamp", "nth highest (zero index)"

    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sheet_name, timestamp_str, nth_zero, *expected_vals = row
        meas_name = sheet_to_name.get(sheet_name)
        if meas_name is None:
            errors.append(f"Row {row_idx}: sheet {sheet_name!r} not found in Nor145 file")
            continue

        log = name_to_log.get(meas_name)
        if log is None:
            errors.append(f"Row {row_idx}: measurement {meas_name!r} not found in parsed data")
            continue

        # Convert 0-indexed nth to 1-indexed
        n = int(nth_zero) + 1

        # Determine pivot from the first spectral column
        first_col = _oracle_col_to_pycoustic(spectral_col_names[0])
        pivot_col = first_col

        # Determine high/low from the metric name
        high = first_col[0].startswith("Lmax")

        result = log.get_nth_high_low(
            n=n,
            count=1,
            pivot_col=pivot_col,
            all_cols=True,
            group_by_date=False,
            high=high,
        )

        if result.empty:
            errors.append(
                f"Row {row_idx}: get_nth_high_low(n={n}, pivot={pivot_col}) "
                f"returned empty for {meas_name!r}"
            )
            continue

        # Compare each spectral value
        for col_name, expected in zip(spectral_col_names, expected_vals):
            if expected is None:
                continue
            pycoustic_col = _oracle_col_to_pycoustic(col_name)
            if pycoustic_col not in result.columns:
                errors.append(
                    f"Row {row_idx}: column {pycoustic_col} not in result for {meas_name!r}"
                )
                continue
            actual = result[pycoustic_col].iloc[0]
            expected_f = float(expected)
            if abs(actual - expected_f) > 0.05:
                errors.append(
                    f"Row {row_idx} ({meas_name}, n={n}, {col_name}): "
                    f"expected {expected_f}, got {actual}"
                )

    wb.close()

    if errors:
        pytest.fail("\n".join(errors[:20]))


def test_nth_high_low_oracle_low(oracle_path, name_to_log, sheet_to_name):
    """Compare get_nth_high_low output against pre-computed nth-lowest Leq values."""
    wb = openpyxl.load_workbook(oracle_path)
    ws = wb["low"]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    spectral_col_names = header[3:]

    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sheet_name, timestamp_str, nth_zero, *expected_vals = row
        meas_name = sheet_to_name.get(sheet_name)
        if meas_name is None:
            errors.append(f"Row {row_idx}: sheet {sheet_name!r} not found in Nor145 file")
            continue

        log = name_to_log.get(meas_name)
        if log is None:
            errors.append(f"Row {row_idx}: measurement {meas_name!r} not found in parsed data")
            continue

        n = int(nth_zero) + 1
        first_col = _oracle_col_to_pycoustic(spectral_col_names[0])
        pivot_col = first_col

        result = log.get_nth_high_low(
            n=n,
            count=1,
            pivot_col=pivot_col,
            all_cols=True,
            group_by_date=False,
            high=False,
        )

        if result.empty:
            errors.append(
                f"Row {row_idx}: get_nth_high_low(n={n}, pivot={pivot_col}, high=False) "
                f"returned empty for {meas_name!r}"
            )
            continue

        for col_name, expected in zip(spectral_col_names, expected_vals):
            if expected is None:
                continue
            pycoustic_col = _oracle_col_to_pycoustic(col_name)
            if pycoustic_col not in result.columns:
                errors.append(
                    f"Row {row_idx}: column {pycoustic_col} not in result for {meas_name!r}"
                )
                continue
            actual = result[pycoustic_col].iloc[0]
            expected_f = float(expected)
            if abs(actual - expected_f) > 0.05:
                errors.append(
                    f"Row {row_idx} ({meas_name}, n={n}, {col_name}): "
                    f"expected {expected_f}, got {actual}"
                )

    wb.close()

    if errors:
        pytest.fail("\n".join(errors[:20]))